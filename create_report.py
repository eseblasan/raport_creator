import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

from utils.url_extractor import is_url


def add_link_to_data(group_name: str, link: str, data) -> None:
    data[group_name].append(link)

def customize_excel(file_path: str) -> None:

    """
    This function cutomize Excel file.
    Creates links, and center data in columns. Doing bold projects names
    """

    wb = load_workbook(file_path)
    ws = wb.active
    center_alignment = Alignment(horizontal="center", vertical="center")

    for row in ws.iter_rows():
        for i, cell in enumerate(row):
            cell.alignment = center_alignment

            # bold project name
            if i == 0:
                cell.font = Font(bold=True)

            # custom links
            # Underlines and switch color
            if cell.value:
                val_str = str(cell.value).strip()

                if is_url(val_str):
                    cell.value = str(i)
                    cell.hyperlink = val_str
                    cell.font = Font(color="0000FF", underline="single")

    # auto-width for text in columns
    for col in ws.columns:
        max_length = 0
        column_letter = get_column_letter(col[0].column)

        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass

        adjusted_width = (max_length + 4)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(file_path)

def create_report(data: dict, file_path: str) -> None:
    """
    This function creates the Excel report uses customize function customize_excel()
    """

    rows = []

    for group_name, links in data.items():
        row = [group_name.upper()] + links
        rows.append(row)

    df = pd.DataFrame(rows)
    df.to_excel(file_path, index=False, header=False)

    customize_excel(file_path)
